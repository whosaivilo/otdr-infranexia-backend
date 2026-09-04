#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import sys
import json
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# === KONSTANTA & KAMUS DATA ===
RX_ONU_BASE = -16.0
DEFAULT_ODC = "ODC DUM FH"
THRESHOLD_VALUE = 7.0614781398215

CABLE_INFO = { 14: "kabel 264", 17: "150m", 18: "kabel 264", 24: "150m", 25: "kabel 264", 34: "TITIK" }
REPAIR_INFO = { 12: "TITIK REPAIR", 17: "TITIK REPAIR", 24: "TITIK REPAIR", 32: "TITIK REPAIR", 34: "ODC" }
PROJECT_ROW4 = { 1: "STO", 2: "ODC DUM FH", 4: "8 km", 5: "Panjang kabel 9,.", 7: "TOTAL NILAI BENDING", 12: "200m", 17: "2", 24: "2", 32: "250m", 34: "kabel 48" }

# === STYLING EXCEL ===
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)

FONT_DATA = Font(size=10, name="Calibri")
FONT_HEADER_WHITE = Font(bold=True, size=10, name="Calibri", color="FFFFFF")
FONT_HEADER_BLACK = Font(bold=True, size=10, name="Calibri", color="000000")
FONT_RED = Font(size=10, name="Calibri", color="FF0000")

FILL_BLACK = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
FILL_RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
FILL_HEADER = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

def safe_float(val, default=0.0):
    if val is None: return default
    if isinstance(val, (int, float)): return float(val)
    try: return float(str(val).replace(",", "."))
    except: return default

def parse_event_value(val):
    if val is None: return None
    if isinstance(val, str):
        sval = val.strip().lower()
        if sval == "end": return "end"
        if sval == "begin": return "begin"
        try: return abs(float(sval.replace(",", ".")))
        except: return None
    if isinstance(val, (int, float)): return abs(float(val))
    return None

def read_input_file(filepath):
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    date_val = ws.cell(row=3, column=2).value
    date_str = str(date_val).strip() if date_val else datetime.now().strftime("%m/%d/%Y %H:%M:%S")

    distance_headers = []
    col_idx = 8
    while True:
        val = ws.cell(row=5, column=col_idx).value
        if val is None: break
        try: distance_headers.append(float(val))
        except: break
        col_idx += 1

    rows_data = []
    row_idx = 6
    while True:
        file_name = ws.cell(row=row_idx, column=2).value
        if not file_name or str(file_name).strip() == "": break

        events = []
        for d_idx in range(len(distance_headers)):
            events.append(parse_event_value(ws.cell(row=row_idx, column=8 + d_idx).value))

        redaman_core = sum(v for v in events if isinstance(v, (int, float)))
        attenuation = safe_float(ws.cell(row=row_idx, column=7).value)
        estimasi_rx = RX_ONU_BASE - redaman_core - attenuation

        rows_data.append({
            "filename": str(file_name).strip(),
            "fiber": str(ws.cell(row=row_idx, column=3).value or "").strip(),
            "wavelength": str(ws.cell(row=row_idx, column=4).value or "").strip(),
            "loss_db": safe_float(ws.cell(row=row_idx, column=5).value),
            "length_km": safe_float(ws.cell(row=row_idx, column=6).value),
            "attenuation": attenuation,
            "estimasi_rx_onu": round(estimasi_rx, 3),
            "redaman_core": round(redaman_core, 3),
            "loss": round(redaman_core, 3),
            "events": events
        })
        row_idx += 1
    wb.close()
    return {"date": date_str, "distance_headers": distance_headers, "rows": rows_data}

def compute_summary(distance_headers, rows_data):
    num_cols = len(distance_headers)
    jumlah_tp, jumlah_bend, total_bend = [], [], []
    for col_idx in range(num_cols):
        tp, jb, tnb = 0, 0, 0.0
        for row in rows_data:
            val = row["events"][col_idx]
            if val == "end": tp += 1
            elif isinstance(val, (int, float)):
                jb += 1
                tnb += val
        jumlah_tp.append(tp)
        jumlah_bend.append(jb)
        total_bend.append(round(distance_headers[col_idx] + tnb, 13))
    return {"tp": jumlah_tp, "bend": jumlah_bend, "total": total_bend}

def create_formatted_excel(output_path, raw_data, summary, threshold):
    wb = Workbook()
    ws = wb.active
    ws.title = "Event Table"

    # 1. Baris Summary
    headers_summary = ["JUMLAH TITIK PUTUS", "JUMLAH BENDING & TIPUS", "TOTAL NILAI BENDING"]
    for i, summary_text in enumerate(headers_summary, start=1):
        cell = ws.cell(row=i, column=9, value=summary_text)
        cell.font, cell.fill, cell.alignment = FONT_HEADER_WHITE, FILL_BLACK, CENTER_ALIGN

        for d_idx in range(len(raw_data["distance_headers"])):
            col = 10 + d_idx
            val = summary[["tp", "bend", "total"][i-1]][d_idx]
            ws.cell(row=i, column=col, value=val).alignment = CENTER_ALIGN

    # 2. Info Proyek (Baris 4 & 5)
    for col, val in PROJECT_ROW4.items():
        cell = ws.cell(row=4, column=col, value=val)
        cell.font, cell.alignment = FONT_HEADER_BLACK, LEFT_ALIGN
        if val == "TOTAL NILAI BENDING":  # PERBAIKAN TYPO DI SINI
            cell.fill, cell.font = FILL_BLACK, FONT_HEADER_WHITE

    for col, val in CABLE_INFO.items():
        cell = ws.cell(row=5, column=col, value=val)
        cell.font, cell.alignment = FONT_DATA, CENTER_ALIGN
        if val in ["150m", "200m", "250m", "TITIK"]:
            cell.fill, cell.font = FILL_RED, FONT_HEADER_WHITE

    # 3. Tanggal, Repair & ODC (Baris 6)
    ws.cell(row=6, column=1, value="Date:").font = FONT_HEADER_BLACK
    ws.cell(row=6, column=2, value=raw_data["date"]).font = FONT_DATA
    for col, val in REPAIR_INFO.items():
        cell = ws.cell(row=6, column=col, value=val)
        cell.font, cell.alignment = FONT_HEADER_BLACK, CENTER_ALIGN

    # 4. Header Tabel
    headers_col = [(2,"File"), (3,"Fiber"), (4,"Wavelength"), (5,"Loss, dB"), (6,"Length, km"), (7,"Attenuation, dB/km"), (8,"ESTIMASI RX ONU"), (9,"REDAMAN / CORE")]
    for col, val in headers_col:
        cell = ws.cell(row=7, column=col, value=val)
        if col in [8, 9]: cell.font, cell.fill = FONT_HEADER_WHITE, FILL_BLACK
        else: cell.font, cell.fill = FONT_HEADER_BLACK, FILL_HEADER
        cell.alignment, cell.border = CENTER_ALIGN, THIN_BORDER

    for d_idx, dist in enumerate(raw_data["distance_headers"]):
        cell = ws.cell(row=7, column=10 + d_idx, value=dist)
        cell.font, cell.fill, cell.alignment, cell.border = FONT_HEADER_BLACK, FILL_HEADER, CENTER_ALIGN, THIN_BORDER

    # Tambahan Header Loss dan Threshold
    loss_col = 10 + len(raw_data["distance_headers"])
    thresh_col = loss_col + 1
    for col, val in [(loss_col, "Loss"), (thresh_col, "Threshold")]:
        cell = ws.cell(row=7, column=col, value=val)
        cell.font, cell.fill, cell.alignment, cell.border = FONT_HEADER_BLACK, FILL_HEADER, CENTER_ALIGN, THIN_BORDER

    # 5. Isi Data (Baris 8 ke bawah)
    current_row = 8
    for row in raw_data["rows"]:
        ws.cell(row=current_row, column=2, value=row["filename"]).font = FONT_DATA
        ws.cell(row=current_row, column=3, value=row["fiber"]).font = FONT_DATA
        ws.cell(row=current_row, column=4, value=row["wavelength"]).font = FONT_DATA
        ws.cell(row=current_row, column=5, value=row["loss_db"]).font = FONT_DATA
        ws.cell(row=current_row, column=6, value=row["length_km"]).font = FONT_DATA
        ws.cell(row=current_row, column=7, value=row["attenuation"]).font = FONT_DATA

        # Sel Estimasi RX ONU (Kuning jika kurang dari -22 dBm)
        cell_rx = ws.cell(row=current_row, column=8, value=row["estimasi_rx_onu"])
        cell_rx.font, cell.alignment = FONT_DATA, CENTER_ALIGN
        if row["estimasi_rx_onu"] < -22.0: cell_rx.fill = FILL_YELLOW

        ws.cell(row=current_row, column=9, value=row["redaman_core"]).font = FONT_DATA

        for d_idx in range(len(raw_data["distance_headers"])):
            val = row["events"][d_idx]
            cell = ws.cell(row=current_row, column=10 + d_idx, value=val)
            cell.alignment = CENTER_ALIGN
            if val == "end":
                cell.font, cell.fill = FONT_RED, FILL_YELLOW
            elif isinstance(val, (int, float)):
                cell.font, cell.fill = FONT_DATA, FILL_YELLOW
            else:
                cell.font = FONT_DATA

        ws.cell(row=current_row, column=loss_col, value=row["loss"]).font = FONT_DATA
        ws.cell(row=current_row, column=thresh_col, value=threshold).font = FONT_DATA

        current_row += 1

    # Lebar Kolom
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    for col in range(10, 12 + len(raw_data["distance_headers"])):
        ws.column_dimensions[get_column_letter(col)].width = 12

    wb.save(output_path)
    wb.close()

def process_file(filepath, output_path, odc_name=None, threshold=None):
    raw_data = read_input_file(filepath)
    summary = compute_summary(raw_data["distance_headers"], raw_data["rows"])
    thr = threshold if threshold is not None else DEFAULT_THRESHOLD
    for row in raw_data["rows"]: row["threshold"] = thr

    create_formatted_excel(output_path, raw_data, summary, thr)

    return {
        "odc": odc_name if odc_name else DEFAULT_ODC,
        "date": raw_data["date"],
        "distance_headers": raw_data["distance_headers"],
        "summary": summary,
        "rows": raw_data["rows"]
    }

if __name__ == "__main__":
    if len(sys.argv) < 3: sys.exit(1)
    try:
        result = process_file(sys.argv[1], sys.argv[2], sys.argv[3] if len(sys.argv) > 3 else None)
        print(json.dumps(result))
    except Exception as e:
        print(json.dumps({"error": str(e)}))
        sys.exit(1)
